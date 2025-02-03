'''
SPDX-License-Identifier: Apache-2.0

Copyright 2023 Cisco Systems, Inc. and its affiliates

Licensed under the Apache License, Version 2.0 (the "License");
you may not use this file except in compliance with the License.
You may obtain a copy of the License at

http://www.apache.org/licenses/LICENSE-2.0

Unless required by applicable law or agreed to in writing, software
distributed under the License is distributed on an "AS IS" BASIS,
WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
See the License for the specific language governing permissions and
limitations under the License.
'''

from pptx import *
import sys, os, re
import numpy as np
import math,shutil
import openpyxl
import tkinter as tk ,tkinter.ttk as ttk,tkinter.filedialog, tkinter.messagebox
import ns_def ,ns_ddx_figure , network_sketcher,network_sketcher_dev

def l2_device_table_sync_with_l3_master(self):
    #print('--- l2_device_table_sync_with_l3_master ---')

    ### get exist L2 table sheet in master file
    ws_l2_name = 'Master_Data_L2'
    ws_l3_name = 'Master_Data_L3'
    excel_maseter_file = self.full_filepath

    original_master_l2_table_array = ns_def.convert_master_to_array(ws_l2_name, excel_maseter_file, '<<L2_TABLE>>')
    original_master_l3_table_array = ns_def.convert_master_to_array(ws_l3_name, excel_maseter_file, '<<L3_TABLE>>')

    del original_master_l2_table_array[:2]
    del original_master_l3_table_array[:2]

    #print(original_master_l2_table_array)
    #print(original_master_l3_table_array)

    for index_l2,tmp_original_master_l2_table_array in enumerate(original_master_l2_table_array):
        tmp_original_master_l2_table_array[1].extend(['','','','','','','',''])
        del tmp_original_master_l2_table_array[1][8:]
        original_master_l2_table_array[index_l2] = tmp_original_master_l2_table_array

    for index_l3,tmp_original_master_l3_table_array in enumerate(original_master_l3_table_array):
        tmp_original_master_l3_table_array[1].extend(['','','','','','','','','',''])
        del tmp_original_master_l3_table_array[1][7:]
        original_master_l3_table_array[index_l3] = tmp_original_master_l3_table_array

    ### remove Master L3 sheet
    ppt_meta_file = excel_maseter_file
    copy_sheet_name = ws_l3_name
    ns_def.remove_excel_sheet(ppt_meta_file, copy_sheet_name)

    ### re-create Master L3 sheet
    #run L3-1-2 in network_sketcher_dev ,  add l3 master sheet
    self.click_value = 'L3-1-2'
    network_sketcher_dev.ns_front_run.click_action(self, 'L3-1-2')

    # remove exist L3/ file
    if os.path.isfile(self.full_filepath.replace('[MASTER]', '[L3_TABLE]')) == True and ns_def.check_file_locked(self.full_filepath.replace('[MASTER]', '[L3_TABLE]')) == False:
        os.remove(self.full_filepath.replace('[MASTER]', '[L3_TABLE]'))

    ### get re-create L3 Master
    recreate_master_l3_table_array = ns_def.convert_master_to_array(ws_l3_name, excel_maseter_file, '<<L3_TABLE>>')
    del recreate_master_l3_table_array[:2]

    for index_l3,tmp_recreate_master_l3_table_array in enumerate(recreate_master_l3_table_array):
        tmp_recreate_master_l3_table_array[1].extend(['','','','','','','','','',''])
        del tmp_recreate_master_l3_table_array[1][7:]
        recreate_master_l3_table_array[index_l3] = tmp_recreate_master_l3_table_array

    ### update original L3 Master values to recreate L3 Master
    for index_l3_2 ,tmp_recreate_master_l3_table_array in enumerate(recreate_master_l3_table_array):
        tmp_recreate_master_l3_table_array[1].append(ns_def.get_if_value(tmp_recreate_master_l3_table_array[1][2]))
        tmp_recreate_master_l3_table_array[1].extend([str(ns_def.split_portname(tmp_recreate_master_l3_table_array[1][2])[0])])
        recreate_master_l3_table_array[index_l3_2] = tmp_recreate_master_l3_table_array

        for tmp_original_master_l3_table_array in original_master_l3_table_array:
            if tmp_recreate_master_l3_table_array[1][1] == tmp_original_master_l3_table_array[1][1] and tmp_recreate_master_l3_table_array[1][2] == tmp_original_master_l3_table_array[1][2]:
                recreate_master_l3_table_array[index_l3_2][1][3] = tmp_original_master_l3_table_array[1][3]
                recreate_master_l3_table_array[index_l3_2][1][4] = tmp_original_master_l3_table_array[1][4]
                recreate_master_l3_table_array[index_l3_2][1][5] = tmp_original_master_l3_table_array[1][5]
                recreate_master_l3_table_array[index_l3_2][1][6] = tmp_original_master_l3_table_array[1][6]

    ### make array for sort
    sort_master_l3_table_array = []
    for tmp_recreate_master_l3_table_array in recreate_master_l3_table_array:
        sort_master_l3_table_array.append(tmp_recreate_master_l3_table_array[1])

    sort_master_l3_table_array = sorted(sort_master_l3_table_array, reverse=False, key=lambda x: (x[0], x[1], x[3], x[8], x[7], x[4]))  # sort l3 table

    last_master_l3_table_array = []
    last_master_l3_table_array.append([1, ['<<L3_TABLE>>']])
    last_master_l3_table_array.append([2, ['Area', 'Device Name', 'L3 IF Name','L3 Instance Name', 'IP Address / Subnet mask (Comma Separated)', '[VPN] Target Device Name (Comma Separated)', '[VPN] Target L3 Port Name (Comma Separated)']])
    for index_l3_3,tmp_sort_master_l3_table_array in enumerate(sort_master_l3_table_array):
        del tmp_sort_master_l3_table_array[-2:]
        last_master_l3_table_array.append([index_l3_3 + 3 ,tmp_sort_master_l3_table_array])

    #print('--- last_master_l3_table_array ---')
    #print(last_master_l3_table_array)

    ### Re-Write L3 sheet
    last_master_l3_table_tuple = ns_def.convert_array_to_tuple(last_master_l3_table_array)
    tmp_ws_name = ws_l3_name
    ppt_meta_file = excel_maseter_file
    clear_section_taple = last_master_l3_table_tuple
    ns_def.clear_section_sheet(tmp_ws_name, ppt_meta_file, clear_section_taple)

    master_excel_meta = last_master_l3_table_tuple
    excel_file_path = excel_maseter_file
    worksheet_name = ws_l3_name
    section_write_to = '<<L3_TABLE>>'
    offset_row = 0
    offset_column = 0
    ns_def.overwrite_excel_meta(master_excel_meta, excel_file_path, worksheet_name, section_write_to, offset_row, offset_column)

def l2_device_port_name_sync_with_l3_master(self):
    #print('--- l2_device_vport_name_sync_with_l3_master ---')
    ws_name = 'Master_Data'
    ws_l2_name = 'Master_Data_L2'
    ws_l3_name = 'Master_Data_L3'
    excel_maseter_file = self.full_filepath
    device_table_file = self.inFileTxt_12_1.get()
    device_l2_table_ws_name = 'L2 Table'
    self.update_l2_table_vport_array = []

    excel_device_table = openpyxl.load_workbook(device_table_file)
    excel_device_table.active = excel_device_table.sheetnames.index(device_l2_table_ws_name)

    # GET Folder and wp name List
    folder_wp_name_array = ns_def.get_folder_wp_array_from_master(ws_name, excel_maseter_file)
    #print('---- folder_wp_name_array ----')
    # print(folder_wp_name_array)


    flag_get_start = False
    for tmp_column in range(1, 100):
        for tmp_row in range(1, 100):
            if '4f81bd' in str(excel_device_table.active.cell(tmp_row, tmp_column).fill.bgColor.value).lower() and 'dce6f1' in str(excel_device_table.active.cell(tmp_row + 1, tmp_column).fill.bgColor.value).lower():
                start_folder_array = tmp_row + 1, tmp_column, excel_device_table.active.cell(tmp_row + 1, tmp_column).value
                for tmp_folder_wp_name_array in folder_wp_name_array[0]:
                    if tmp_folder_wp_name_array == start_folder_array[2]:
                        flag_get_start = True
                        break
        if flag_get_start == True:
            break
    #print('start_folder_array')
    #print(start_folder_array)


    # convert from device table to array and convert to tuple
    excel_ws_name = 'L2 Table'
    excel_file = device_table_file
    start_row = start_folder_array[0]
    self.device_l2_table_array = ns_def.convert_excel_to_array(excel_ws_name, excel_file, start_row)

    update_device_l2_table_array = []
    current_folder_name = '__dummy__'
    current_shape_name = '__dummy__'
    for tmp_device_l2_table_array in self.device_l2_table_array:
        tmp_device_l2_table_array[1].extend(['','','','','','','','',''])
        del tmp_device_l2_table_array[1][8:]
        tmp_device_l2_table_array[1][2] = ''
        tmp_device_l2_table_array[1][4] = ''

        if tmp_device_l2_table_array[1][0] != '':
            current_folder_name = tmp_device_l2_table_array[1][0]
        elif tmp_device_l2_table_array[1][0] == '':
            tmp_device_l2_table_array[1][0] = current_folder_name

        if tmp_device_l2_table_array[1][1] != '':
            current_shape_name = tmp_device_l2_table_array[1][1]
        elif tmp_device_l2_table_array[1][1] == '':
            tmp_device_l2_table_array[1][1] = current_shape_name

        update_device_l2_table_array.append(tmp_device_l2_table_array[1])

        if '\n' in tmp_device_l2_table_array[1][5]:
            tmp_update_if_name = tmp_device_l2_table_array[1][5]
            tmp_update_if_array = tmp_update_if_name.split('\n')
            self.update_l2_table_vport_array.append([tmp_device_l2_table_array[1][1],tmp_update_if_array[0],tmp_update_if_array[1]])


    #print('--- update_device_l2_table_array. ---')
    #print(update_device_l2_table_array)

    #print('--- self.update_l2_table_vport_array ---')
    #print(self.update_l2_table_vport_array)

    ### make overwrite tuple

    ws_l2_name = 'Master_Data_L2'
    ws_l3_name = 'Master_Data_L3'
    excel_maseter_file = self.full_filepath

    master_l2_table_array = ns_def.convert_master_to_array(ws_l2_name, excel_maseter_file, '<<L2_TABLE>>')
    master_l3_table_array = ns_def.convert_master_to_array(ws_l3_name, excel_maseter_file, '<<L3_TABLE>>')

    for index_l2,tmp_master_l2_table_array in enumerate(master_l2_table_array):
        tmp_master_l2_table_array[1].extend(['','','','','','','',''])
        del tmp_master_l2_table_array[1][8:]
        master_l2_table_array[index_l2] = tmp_master_l2_table_array

    for index_l3,tmp_master_l3_table_array in enumerate(master_l3_table_array):
        tmp_master_l3_table_array[1].extend(['','','','','','','','','','',])
        del tmp_master_l3_table_array[1][7:]
        master_l3_table_array[index_l3] = tmp_master_l3_table_array

    #convert to tuple
    master_l2_table_tuple = ns_def.convert_array_to_tuple(master_l2_table_array)
    master_l3_table_tuple = ns_def.convert_array_to_tuple(master_l3_table_array)

    overwrite_master_l2_table_tuple = {}
    overwrite_master_l3_table_tuple = {}

    for tmp_master_l2_table_tuple in master_l2_table_tuple:
        if tmp_master_l2_table_tuple[1] == 2 and tmp_master_l2_table_tuple[0] != 1 and tmp_master_l2_table_tuple[0] != 2:
            for tmp_update_l2_table_vport_array in self.update_l2_table_vport_array:
                if tmp_update_l2_table_vport_array[0] == master_l2_table_tuple[tmp_master_l2_table_tuple[0],tmp_master_l2_table_tuple[1]] and tmp_update_l2_table_vport_array[1] == master_l2_table_tuple[tmp_master_l2_table_tuple[0],6]:
                    overwrite_master_l2_table_tuple[tmp_master_l2_table_tuple[0],6] = tmp_update_l2_table_vport_array[2]
                    break

    for tmp_master_l3_table_tuple in master_l3_table_tuple:
        if tmp_master_l3_table_tuple[1] == 2 and tmp_master_l3_table_tuple[0] != 1 and tmp_master_l3_table_tuple[0] != 2:
            for tmp_update_l3_table_vport_array in self.update_l2_table_vport_array:
                if tmp_update_l3_table_vport_array[0] == master_l3_table_tuple[tmp_master_l3_table_tuple[0],tmp_master_l3_table_tuple[1]] and tmp_update_l3_table_vport_array[1] == master_l3_table_tuple[tmp_master_l3_table_tuple[0],3]:
                    overwrite_master_l3_table_tuple[tmp_master_l3_table_tuple[0],3] = tmp_update_l3_table_vport_array[2]
                    break

    #print('--- overwrite_master_l2_table_tuple ,overwrite_master_l3_table_tuple---')
    #print(overwrite_master_l2_table_tuple,overwrite_master_l3_table_tuple)

    ### write master l2/l3
    master_excel_meta = overwrite_master_l2_table_tuple
    excel_file_path = excel_maseter_file
    worksheet_name = ws_l2_name
    section_write_to = '<<L2_TABLE>>'
    offset_row = 0
    offset_column = 0
    ns_def.overwrite_excel_meta(master_excel_meta, excel_file_path, worksheet_name, section_write_to,offset_row, offset_column)

    master_excel_meta = overwrite_master_l3_table_tuple
    excel_file_path = excel_maseter_file
    worksheet_name = ws_l3_name
    section_write_to = '<<L3_TABLE>>'
    offset_row = 0
    offset_column = 0
    ns_def.overwrite_excel_meta(master_excel_meta, excel_file_path, worksheet_name, section_write_to,offset_row, offset_column)


def l1_device_port_name_sync_with_l2l3_master(self):
    #print('--- l1_device_port_name_sync_with_l2l3_master ---')
    #print(self.update_port_num_array)

    ws_l2_name = 'Master_Data_L2'
    ws_l3_name = 'Master_Data_L3'
    excel_maseter_file = self.full_filepath

    master_l2_table_array = ns_def.convert_master_to_array(ws_l2_name, excel_maseter_file, '<<L2_TABLE>>')
    master_l3_table_array = ns_def.convert_master_to_array(ws_l3_name, excel_maseter_file, '<<L3_TABLE>>')

    #convert to tuple
    master_l2_table_tuple = ns_def.convert_array_to_tuple(master_l2_table_array)
    master_l3_table_tuple = ns_def.convert_array_to_tuple(master_l3_table_array)

    overwrite_l2_tuple = {}
    overwrite_l3_tuple = {}

    ### make change l2/l3 tuple for overwrire
    for tmp_master_l2_table_tuple in master_l2_table_tuple:
        if tmp_master_l2_table_tuple[1] == 2 and tmp_master_l2_table_tuple[0] != 1 and tmp_master_l2_table_tuple[0] != 2:
            tmp_device_name = master_l2_table_tuple[tmp_master_l2_table_tuple[0],tmp_master_l2_table_tuple[1]]
            for tmp_update_port_num_array in self.update_port_num_array:
                if tmp_update_port_num_array[0] == tmp_device_name and tmp_update_port_num_array[1] == master_l2_table_tuple[tmp_master_l2_table_tuple[0],4]:
                    overwrite_l2_tuple[tmp_master_l2_table_tuple[0],4] = tmp_update_port_num_array[2]
                    break

    for tmp_master_l3_table_tuple in master_l3_table_tuple:
        if tmp_master_l3_table_tuple[1] == 2 and tmp_master_l3_table_tuple[0] != 1 and tmp_master_l3_table_tuple[0] != 2:
            tmp_device_name = master_l3_table_tuple[tmp_master_l3_table_tuple[0], tmp_master_l3_table_tuple[1]]
            for tmp_update_port_num_array in self.update_port_num_array:
                if tmp_update_port_num_array[0] == tmp_device_name and tmp_update_port_num_array[1] == master_l3_table_tuple[tmp_master_l3_table_tuple[0], 3]:
                    overwrite_l3_tuple[tmp_master_l3_table_tuple[0], 3] = tmp_update_port_num_array[2]
                    break

    #print('--- overwrite_l2_tuple ,overwrite_l3_tuple---')
    #print(overwrite_l2_tuple,overwrite_l3_tuple)

    ### write master l2/l3
    master_excel_meta = overwrite_l2_tuple
    excel_file_path = excel_maseter_file
    worksheet_name = ws_l2_name
    section_write_to = '<<L2_TABLE>>'
    offset_row = 0
    offset_column = 0
    ns_def.overwrite_excel_meta(master_excel_meta, excel_file_path, worksheet_name, section_write_to,offset_row, offset_column)

    master_excel_meta = overwrite_l3_tuple
    excel_file_path = excel_maseter_file
    worksheet_name = ws_l3_name
    section_write_to = '<<L3_TABLE>>'
    offset_row = 0
    offset_column = 0
    ns_def.overwrite_excel_meta(master_excel_meta, excel_file_path, worksheet_name, section_write_to,offset_row, offset_column)


def l1_sketch_device_name_sync_with_l2l3_master(self):
    ws_l2_name = 'Master_Data_L2'
    ws_l3_name = 'Master_Data_L3'
    excel_maseter_file = self.full_filepath

    master_l2_table_array = ns_def.convert_master_to_array(ws_l2_name, excel_maseter_file, '<<L2_TABLE>>')
    master_l3_table_array = ns_def.convert_master_to_array(ws_l3_name, excel_maseter_file, '<<L3_TABLE>>')

    #convert to tuple
    master_l2_table_tuple = ns_def.convert_array_to_tuple(master_l2_table_array)
    master_l3_table_tuple = ns_def.convert_array_to_tuple(master_l3_table_array)

    #clear section
    tmp_ws_name = ws_l2_name
    ppt_meta_file = excel_maseter_file
    clear_section_taple = master_l2_table_tuple
    ns_def.clear_section_sheet(tmp_ws_name, ppt_meta_file, clear_section_taple)

    tmp_ws_name = ws_l3_name
    ppt_meta_file = excel_maseter_file
    clear_section_taple = master_l3_table_tuple
    ns_def.clear_section_sheet(tmp_ws_name, ppt_meta_file, clear_section_taple)


    #update device name in Master_Data_L2 sheet
    for index_l2_1 , tmp_master_l2_table_array in  enumerate(master_l2_table_array):
        for index_l2_2, tmp_tmp_master_l2_table_array in enumerate(tmp_master_l2_table_array[1]):
            if index_l2_2 == 0 or index_l2_2 == 1:
                for tmp_updated_name_array in self.updated_name_array:
                    if tmp_updated_name_array[0] == tmp_tmp_master_l2_table_array:
                        master_l2_table_array[index_l2_1][1][index_l2_2] = tmp_updated_name_array[1]
                        #print('# update# ', master_l2_table_array[index_l2_1][1][index_l2_2] , tmp_tmp_master_l2_table_array)


    #update device name in Master_Data_L3 sheet
    for index_l3_1 , tmp_master_l3_table_array in  enumerate(master_l3_table_array):
        for index_l3_2, tmp_tmp_master_l3_table_array in enumerate(tmp_master_l3_table_array[1]):
            if index_l3_2 == 0 or index_l3_2 == 1:
                for tmp_updated_name_array in self.updated_name_array:
                    if tmp_updated_name_array[0] == tmp_tmp_master_l3_table_array:
                        master_l3_table_array[index_l3_1][1][index_l3_2] = tmp_updated_name_array[1]
                        #print('# update# ', master_l3_table_array[index_l3_1][1][index_l3_2], tmp_tmp_master_l3_table_array)

    #convert to tuple
    master_l2_table_tuple = ns_def.convert_array_to_tuple(master_l2_table_array)
    master_l3_table_tuple = ns_def.convert_array_to_tuple(master_l3_table_array)

    #write updated array in Master_Data_L2/L3 sheet
    master_excel_meta = master_l2_table_tuple
    excel_file_path = excel_maseter_file
    worksheet_name = ws_l2_name
    section_write_to = '<<L2_TABLE>>'
    offset_row =  0
    offset_column = 0
    ns_def.overwrite_excel_meta(master_excel_meta, excel_file_path, worksheet_name, section_write_to,offset_row, offset_column)

    master_excel_meta = master_l3_table_tuple
    excel_file_path = excel_maseter_file
    worksheet_name = ws_l3_name
    section_write_to = '<<L3_TABLE>>'
    offset_row =  0
    offset_column = 0
    ns_def.overwrite_excel_meta(master_excel_meta, excel_file_path, worksheet_name, section_write_to,offset_row, offset_column)


def l1_master_device_and_line_sync_with_l2l3_master(self):
    #print('l1_master_device_and_line_sync_with_l2l3_master')

    ### get exist L2 table sheet in master file
    ws_l2_name = 'Master_Data_L2'
    ws_l3_name = 'Master_Data_L3'
    excel_maseter_file = self.full_filepath

    original_master_l2_table_array = ns_def.convert_master_to_array(ws_l2_name, excel_maseter_file, '<<L2_TABLE>>')
    original_master_l3_table_array = ns_def.convert_master_to_array(ws_l3_name, excel_maseter_file, '<<L3_TABLE>>')

    del original_master_l2_table_array[:2]
    del original_master_l3_table_array[:2]

    #print(original_master_l2_table_array)
    #print(original_master_l3_table_array)

    for index_l2,tmp_original_master_l2_table_array in enumerate(original_master_l2_table_array):
        tmp_original_master_l2_table_array[1].extend(['','','','','','','',''])
        del tmp_original_master_l2_table_array[1][8:]
        original_master_l2_table_array[index_l2] = tmp_original_master_l2_table_array

    for index_l3,tmp_original_master_l3_table_array in enumerate(original_master_l3_table_array):
        tmp_original_master_l3_table_array[1].extend(['','','','','','','',''])
        del tmp_original_master_l3_table_array[1][7:] # bug fix ns-010 at 2.3.2(b)
        original_master_l3_table_array[index_l3] = tmp_original_master_l3_table_array

    #print(original_master_l2_table_array)
    #print(original_master_l3_table_array)

    '''
    sync with Mater sheet L2
    '''
    ### Master sheet L2 re create from L1 sheet
    ns_def.remove_excel_sheet(excel_maseter_file, ws_l2_name)
    ns_def.remove_excel_sheet(excel_maseter_file, ws_l3_name)

    click_value = 'self.sub1_1_button_1'
    push_array = ['','']
    network_sketcher.ns_front_run.click_action_sub1_1(self, click_value, push_array)

    ### get re-created Master sheet L2/L3
    recreate_master_l2_table_array = ns_def.convert_master_to_array(ws_l2_name, excel_maseter_file, '<<L2_TABLE>>')

    del recreate_master_l2_table_array[:2]

    for index_l2,tmp_recreate_master_l2_table_array in enumerate(recreate_master_l2_table_array):
        tmp_recreate_master_l2_table_array[1].extend(['','','','','','','',''])
        del tmp_recreate_master_l2_table_array[1][8:]
        recreate_master_l2_table_array[index_l2] = tmp_recreate_master_l2_table_array

    #print(recreate_master_l2_table_array)

    ### Update Master sheet L2
    new_master_l2_table_array = []
    for tmp_recreate_master_l2_table_array in recreate_master_l2_table_array:
        flag_match_line_l2 = False
        for tmp_original_master_l2_table_array in original_master_l2_table_array:
            if tmp_original_master_l2_table_array[1][3] != '':
                if tmp_original_master_l2_table_array[1][0] == tmp_recreate_master_l2_table_array[1][0] and tmp_original_master_l2_table_array[1][1] == tmp_recreate_master_l2_table_array[1][1] and tmp_original_master_l2_table_array[1][3] == tmp_recreate_master_l2_table_array[1][3]:
                    tmp_original_master_l2_table_array[1].append(ns_def.get_if_value(tmp_original_master_l2_table_array[1][3]))
                    tmp_original_master_l2_table_array[1].extend([str(ns_def.split_portname(tmp_original_master_l2_table_array[1][3])[0])])
                    new_master_l2_table_array.append(tmp_original_master_l2_table_array[1])
                    flag_match_line_l2 = True
                    break
        if flag_match_line_l2 == False:
            tmp_recreate_master_l2_table_array[1].append(ns_def.get_if_value(tmp_recreate_master_l2_table_array[1][3]))
            tmp_recreate_master_l2_table_array[1].extend([str(ns_def.split_portname(tmp_recreate_master_l2_table_array[1][3])[0])])
            new_master_l2_table_array.append(tmp_recreate_master_l2_table_array[1])

    for tmp_original_master_l2_table_array in original_master_l2_table_array:
        if tmp_original_master_l2_table_array[1][3] == '':
            tmp_original_master_l2_table_array[1].append(0)
            tmp_original_master_l2_table_array[1].append('')
            new_master_l2_table_array.append( tmp_original_master_l2_table_array[1])

    new_master_l2_table_array = sorted(new_master_l2_table_array , reverse=False, key=lambda x: (x[0], x[1], x[9], x[8], x[5], x[7], x[6]))  # sort l2 table

    last_l2_table_array = []
    last_l2_table_array.append([1, ['<<L2_TABLE>>']])
    last_l2_table_array.append([2, ['Area', 'Device Name', 'Port Mode', 'Port Name', 'Virtual Port Mode', 'Virtual Port Name', 'Connected L2 Segment Name(Comma Separated)', 'L2 Name directly received by L3 Virtual Port (Comma Separated)']])
    for index_l2_num, tmp_new_master_l2_table_array in enumerate(new_master_l2_table_array):
        del tmp_new_master_l2_table_array[-2:]
        last_l2_table_array.append([index_l2_num + 3 ,tmp_new_master_l2_table_array])
        #print([index_l2_num + 3 ,tmp_new_master_l2_table_array])


    ### Re-Write L2 sheet
    last_master_l2_table_tuple  = ns_def.convert_array_to_tuple(last_l2_table_array )
    tmp_ws_name   = ws_l2_name
    ppt_meta_file   = excel_maseter_file
    clear_section_taple   = last_master_l2_table_tuple
    ns_def.clear_section_sheet(tmp_ws_name, ppt_meta_file, clear_section_taple)

    master_excel_meta  = last_master_l2_table_tuple
    excel_file_path  = excel_maseter_file
    worksheet_name  = ws_l2_name
    section_write_to  =  '<<L2_TABLE>>'
    offset_row  = 0
    offset_column  = 0
    ns_def.overwrite_excel_meta(master_excel_meta, excel_file_path, worksheet_name, section_write_to,offset_row, offset_column)


    '''
    sync with Mater sheet L3
    '''
    ### Master sheet L3 re create from L1L2 sheet
    ns_def.remove_excel_sheet(excel_maseter_file, ws_l3_name)

    ### run L3-1-2 in network_sketcher_dev ,  add l3 master sheet
    self.click_value = 'L3-1-2'
    self.inFileTxt_L3_1_1.delete(0, tkinter.END)
    self.inFileTxt_L3_1_1.insert(tk.END, excel_maseter_file)
    network_sketcher_dev.ns_front_run.click_action(self, 'L3-1-2')

    ### get re-created Master sheet L2/L3
    recreate_master_l3_table_array = ns_def.convert_master_to_array(ws_l3_name, excel_maseter_file, '<<L3_TABLE>>')

    del recreate_master_l3_table_array[:2]

    for index_l3,tmp_recreate_master_l3_table_array in enumerate(recreate_master_l3_table_array):
        tmp_recreate_master_l3_table_array[1].extend(['','','','','','','',''])
        del tmp_recreate_master_l3_table_array[1][5:]
        recreate_master_l3_table_array[index_l3] = tmp_recreate_master_l3_table_array

    #print(recreate_master_l3_table_array)

    ### Update Master sheet L3
    new_master_l3_table_array = []
    for tmp_recreate_master_l3_table_array in recreate_master_l3_table_array:
        flag_match_line_l3 = False
        for tmp_original_master_l3_table_array in original_master_l3_table_array:
            if tmp_original_master_l3_table_array[1][0] == tmp_recreate_master_l3_table_array[1][0] and tmp_original_master_l3_table_array[1][1] == tmp_recreate_master_l3_table_array[1][1] and tmp_original_master_l3_table_array[1][2] == tmp_recreate_master_l3_table_array[1][2]:
                tmp_original_master_l3_table_array[1].append(ns_def.get_if_value(tmp_original_master_l3_table_array[1][2]))
                tmp_original_master_l3_table_array[1].extend([str(ns_def.split_portname(tmp_original_master_l3_table_array[1][2])[0])])
                new_master_l3_table_array.append(tmp_original_master_l3_table_array[1])
                flag_match_line_l3 = True
                break

        if flag_match_line_l3 == False:

            tmp_recreate_master_l3_table_array[1].append(ns_def.get_if_value(tmp_recreate_master_l3_table_array[1][2]))
            tmp_recreate_master_l3_table_array[1].extend([str(ns_def.split_portname(tmp_recreate_master_l3_table_array[1][2])[0])])
            new_master_l3_table_array.append(tmp_recreate_master_l3_table_array[1])

    new_master_l3_table_array = sorted(new_master_l3_table_array, reverse=False, key=lambda x: (x[0], x[1], x[3], x[6], x[5], x[4]))  # sort l3 table

    last_l3_table_array = []
    last_l3_table_array.append([1, ['<<L3_TABLE>>']])
    last_l3_table_array.append([2, ['Area', 'Device Name', 'L3 IF Name','L3 Instance Name', 'IP Address / Subnet mask (Comma Separated)']])
    for index_l3_num, tmp_new_master_l3_table_array in enumerate(new_master_l3_table_array):
        del tmp_new_master_l3_table_array[-2:]
        last_l3_table_array.append([index_l3_num + 3, tmp_new_master_l3_table_array])
        #print([index_l3_num + 2 ,tmp_new_master_l3_table_array])


    ### Re-Write L3 sheet
    last_master_l3_table_tuple = ns_def.convert_array_to_tuple(last_l3_table_array)
    tmp_ws_name = ws_l3_name
    ppt_meta_file = excel_maseter_file
    clear_section_taple = last_master_l3_table_tuple
    ns_def.clear_section_sheet(tmp_ws_name, ppt_meta_file, clear_section_taple)

    master_excel_meta = last_master_l3_table_tuple
    excel_file_path = excel_maseter_file
    worksheet_name = ws_l3_name
    section_write_to = '<<L3_TABLE>>'
    offset_row = 0
    offset_column = 0
    ns_def.overwrite_excel_meta(master_excel_meta, excel_file_path, worksheet_name, section_write_to, offset_row, offset_column)

    # remove exist L3/ file
    if os.path.isfile(self.inFileTxt_L2_1_1.get().replace('[MASTER]', '[L3_TABLE]')) == True:
        os.remove(self.inFileTxt_L2_1_1.get().replace('[MASTER]', '[L3_TABLE]'))






